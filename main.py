import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill

# Function to get building data from user input
def get_building_data():
    buildings = []

    while True:
        building_name = int(input("동을 입력하세요 (종료하려면 0 입력): "))
        if building_name == 0:
            break
        while True:
            building_type = input("타입을 입력하세요 (해당 동 종료 시, end 입력): ")
            if building_type == "end":
                break
            room_num = int(input("라인을 입력하세요: "))
            highest_floor = int(input("최상층을 입력하세요: "))
            lowest_floor = int(input("최하층을 입력하세요: "))

            # Create list of floors
            floors = list(range(lowest_floor, highest_floor + 1))

            # Append the building information
            building = {
                "동": building_name,
                "타입": f"{building_type}",
                "층": floors,
                "호": room_num,
                "최저층": lowest_floor
            }
            buildings.append(building)

    return buildings


# Function to create nested data structure from building input
def create_nested_data(buildings):
    nested_data = {}

    for building in buildings:
        building_name = building["동"]
        building_type = building["타입"]
        room_num = building["호"]

        if building_name not in nested_data:
            nested_data[building_name] = {}

        if room_num not in nested_data[building_name]:
            nested_data[building_name][room_num] = {}

        if building_type not in nested_data[building_name][room_num]:
            nested_data[building_name][room_num][building_type] = []

        for floor in building["층"]:
            room_number = floor * 100 + room_num  # Generate room number by combining floor and room number
            nested_data[building_name][room_num][building_type].append(room_number)

    return nested_data



from openpyxl.styles import PatternFill

def create_grid_layout(nested_data, output_filename):
    # Create a new workbook and sheet using openpyxl
    wb = openpyxl.Workbook()
    set = wb.active
    set.title = "Grid"

    title = input("현장명을 입력하세요: ")
    set.cell(row=2, column=4).value = title

    # 각 동의 최하층 및 최상층 계산
    min_floor = 1
    max_floor = 0

    buildings_info = {}

    for building_name in nested_data:
        for line in nested_data[building_name]:
            for building_type in nested_data[building_name][line]:
                for room_number in nested_data[building_name][line][building_type]:
                    floor = room_number // 100
                    if building_name not in buildings_info:
                        buildings_info[building_name] = {'lowest_floor': floor, 'highest_floor': floor}
                    else:
                        buildings_info[building_name]['lowest_floor'] = min(buildings_info[building_name]['lowest_floor'], floor)
                        buildings_info[building_name]['highest_floor'] = max(buildings_info[building_name]['highest_floor'], floor)

    # 최하층과 최상층 계산
    for building, info in buildings_info.items():
        min_floor = min(min_floor, info['lowest_floor'])
        max_floor = max(max_floor, info['highest_floor'])

    # 현재 층 정보를 내림차순으로 정렬하여 1열에 작성
    current_row = 4
    val_list = range(max_floor, min_floor - 1, -1)  # Floors in descending order

    for i, val in enumerate(val_list):
        set.cell(row=current_row, column=1, value=val)
        current_row += 1

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    set.cell(row=current_row + 2, column=1, value="타입")
    set.cell(row=current_row + 3, column=1, value="라인")
    set.cell(row=current_row + 4, column=1, value="동")

    current_col = 3
    max_floor += 3

    # Define colors for different floor attributes
    fill_lowest = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow for 최하층
    fill_penthouse = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red for 피트층
    fill_standard = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green for 기준층

    # 동, 라인 번호 우선으로 정렬하여 출력 (라인 순서만 고려)
    for building_name in nested_data.keys():  # 동 이름으로 정렬
        save_col = current_col

        for line in nested_data[building_name].keys():  # 라인 번호로 정렬
            for building_type, rooms in nested_data[building_name][line].items():  # 타입별 방 출력
                set.cell(row=max_floor + 4, column=current_col, value=line).alignment = Alignment(horizontal='center')
                set.cell(row=max_floor + 4, column=current_col).border = thin_border

                lowest_floor = rooms[0] // 100

                # 방 번호를 층 정보를 기준으로 출력
                for room_number in rooms:
                    floor = room_number // 100  # 층 정보
                    room_row = max_floor - floor + 1  # 층에 따른 행 번호 조정
                    if floor == 1:
                        fill = fill_lowest
                    elif floor == lowest_floor :
                        fill = fill_penthouse
                    else:
                        fill = fill_standard

                    set.cell(row=room_row, column=current_col, value=room_number).alignment = Alignment(
                        horizontal='right')
                    set.cell(row=room_row, column=current_col).border = thin_border
                    set.cell(row=room_row, column=current_col).fill = fill

                # 타입 출력 (라인별로 출력하도록 수정)
                set.cell(row=max_floor + 3, column=current_col, value=building_type).alignment = Alignment(
                    horizontal='center')
                set.cell(row=max_floor + 3, column=current_col).border = thin_border

                set.cell(row=max_floor + 5, column=current_col, value=building_name).alignment = Alignment('center')
                set.cell(row=max_floor + 5, column=current_col, value=building_name).border = thin_border

                current_col += 1

        # 한 동의 출력이 끝나면 merge
        set.merge_cells(start_row=max_floor + 5, start_column=save_col, end_row=max_floor + 5,
                        end_column=current_col - 1)
        set.cell(row=max_floor + 5, column=save_col).border = thin_border

        current_col += 2

    wb.save(output_filename)
    print(f"Final layout saved to {output_filename}")


def save_to_excel(nested_data, filename):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Column headers
    ws.cell(row=1, column=1, value="동")
    ws.cell(row=1, column=2, value="라인")
    ws.cell(row=1, column=3, value="타입")
    ws.cell(row=1, column=4, value="호수")
    ws.cell(row=1, column=5, value="층 속성")

    current_row = 2

    for building_name in sorted(nested_data.keys()):  # 동 이름으로 정렬
        for line in sorted(nested_data[building_name].keys()):  # 라인 번호로 정렬
            for building_type in nested_data[building_name][line]:  # 타입을 구분하여 처리
                for room_number in nested_data[building_name][line][building_type]:  # 방 번호 리스트에서 방 번호 가져오기
                    lowest_floor = room_number // 100  # 호수의 첫 자리는 층수

                    # 층 속성 결정
                    if room_number // 100 == 1:
                        floor_attribute = "최하층"
                    elif room_number // 100 == lowest_floor:
                        floor_attribute = "피트층"
                    else:
                        floor_attribute = "기준층"

                    # 데이터를 엑셀에 기록
                    ws.cell(row=current_row, column=1, value=building_name)
                    ws.cell(row=current_row, column=2, value=line)
                    ws.cell(row=current_row, column=3, value=building_type)
                    ws.cell(row=current_row, column=4, value=room_number)
                    ws.cell(row=current_row, column=5, value=floor_attribute)

                    current_row += 1

    # 파일 저장
    wb.save(filename)
    print(f"Data saved to {filename}")


# Main function to integrate all parts
def main():
    # Step 1: Get building data from user
    buildings = get_building_data()

    # Step 2: Create nested data
    nested_data = create_nested_data(buildings)

    # Step 3: Save to intermediate Excel file
    intermediate_file = "building_room_data.xlsx"
    save_to_excel(nested_data, intermediate_file)

    # Step 4: Create grid layout and save final Excel file
    output_file = input("저장하고자하는 최종 파일명을 입력하세요 : ")
    output_file = "save_to_excel/" + output_file + ".xlsx"
    print(nested_data)
    create_grid_layout(nested_data, output_file)


if __name__ == "__main__":
    main()
